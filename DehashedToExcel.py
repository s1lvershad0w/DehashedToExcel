#Imports
import json
import sys
import requests
import argparse
import pandas as pd
import pyfiglet
import jpype
import asposecells
jpype.startJVM()
from asposecells.api import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, NamedStyle, Alignment, Font, Border, Side

#Functions
def check_api_auth_success(dehashed_json_raw):
    check_success = json.loads(dehashed_json_raw)
    if check_success.get('success') == False:
        sys.exit('[-] API Authentication Failure.')
    else:
        pass

def jsonify_data(json_raw_data):
    json_data = json.loads(json_raw_data)
    entries = json_data['entries']
    return entries

def check_data_returned(entries):
    try:
        for x in entries:
            pass
    except TypeError:
        sys.exit('[-] No data returned. Probably error in syntax.')

def query_dehashed_domain(domain, email, key):
    headers = {'Accept': 'application/json',}
    params = (('query', 'domain:' + domain),)
    dehashed_json_raw = requests.get('https://api.dehashed.com/search',
                            headers=headers,
                            params=params,
                            auth=(email, key)).text
    check_api_auth_success(dehashed_json_raw)
    dehashed_json = jsonify_data(dehashed_json_raw)
    return dehashed_json

#Main
prebanner = pyfiglet.figlet_format("DehashedToExcel")
banner = prebanner + "\n-- @FireStone65 -- \n\n"
print(banner)

parser = argparse.ArgumentParser(description='[+] DehashedToExcel.py is a python-based parser for Dehashed to aid OSINT reporting')
parser.add_argument('-d', type=str, required=True, help='Target Domain')
parser.add_argument('-u', type=str, required=True, help='Email ID')
parser.add_argument('-k', type=str, required=True, help='Dehashed API Key')
args = parser.parse_args()

email = args.u
key = args.k
domain = args.d
input_file = domain + '.json'
output_file = domain + '.xlsx'

print('[+] Querying Dehashed for all entries under domain: ' + args.d)
response = query_dehashed_domain(domain, email, key)
check_data_returned(response)
jsonString = json.dumps(response)
print('Raw response received')

# Writing to File
with open(input_file, "w") as outfile:
    outfile.write(jsonString)

print('[+] Raw response saved to ' + input_file)

#Parse JSON to XLSX
with open(input_file, "r") as json_file:
    data = json.load(json_file)

df = pd.DataFrame(data)
df.to_excel(output_file)

#Change Formatting of First Row
wb = load_workbook(output_file)
ws = wb['Sheet1']
ws.delete_cols(1, 2)

named_cols = ['Email', 'IP Address', 'Username', 'Password', 'Hashed Password', 'Name', 'VIN', 'Address']
for index in range(len(named_cols)):
    cell = ws.cell(row = 1, column = index + 1)
    if cell.value is not None:
        cell.value = named_cols[index]

custom_font = Font(name='Calibri', size=11, color='000000', bold=True)
for cell in ws["1:1"]:
    cell.font = custom_font

greyFill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
for cell in ws['1:1']:
    cell.fill = greyFill

custom_Alignment = Alignment(horizontal='left', vertical='bottom', textRotation=0, wrapText=None)
for cell in ws['1:1']:
    cell.alignment = custom_Alignment
wb.save(output_file)

#Autofit Contents of First Column
wb = Workbook(output_file)
worksheet = wb.getWorksheets().get(0)
total_cols = worksheet.getCells().getMaxDataColumn()
for index in range(0, total_cols + 1):
        worksheet.autoFitColumn(index);
wb.save(output_file)

#Delete asposecell's newly added Evaluation Trial worksheet
wb = load_workbook(output_file)
wb['Sheet1'].title = 'Exposed Data in Breaches'
del wb['Evaluation Warning']
jpype.shutdownJVM()

wb.save(output_file)
print('[+] Excel saved to ' + output_file)
